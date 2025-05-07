
import streamlit as st
import pandas as pd
from openpyxl import Workbook
import re

st.title("📊 Excel Generator App")

# Upload files
report_file = st.file_uploader("Upload report.xlsx", type="xlsx")
statement_file = st.file_uploader("Upload statement.xlsx", type="xlsx")

if report_file and statement_file:
    st.success("✅ Files uploaded!")
    report_df = pd.read_excel(report_file)
    statement_df = pd.read_excel(statement_file)

    st.subheader("Report Preview")
    st.dataframe(report_df.head())

    st.subheader("Statement Preview")
    st.dataframe(statement_df.head())

    if st.button("Generate Final File"):
        purchases_df = pd.read_excel(report_file, sheet_name='Grid')
        bank_df = pd.read_excel(statement_file)

        # Helper function
        def extract_id(text):
            digits = ''.join(re.findall(r'\d', str(text)))
            return digits[:11] if len(digits) >= 11 else (digits[:9] if len(digits) >= 9 else "")

        purchases_df['დასახელება'] = purchases_df['გამყიდველი'].astype(str).apply(lambda x: re.sub(r'^\(\d+\)\s*', '', x).strip())
        purchases_df['საიდენტიფიკაციო კოდი'] = purchases_df['გამყიდველი'].apply(extract_id)
        bank_df['P'] = bank_df.iloc[:, 15].astype(str).str.strip()
        bank_df['Amount'] = pd.to_numeric(bank_df.iloc[:, 3], errors='coerce').fillna(0)

        wb = Workbook()
        wb.remove(wb.active)

        ws1 = wb.create_sheet(title="ანგარიშფაქტურები კომპანიით")
        ws1.append(['დასახელება', 'საიდენტიფიკაციო კოდი', 'ანგარიშფაქტურის №', 'ანგარიშფაქტურის თანხა', 'ჩარიცხული თანხა'])

        for company_id, group in purchases_df.groupby('საიდენტიფიკაციო კოდი'):
            company_name = group['დასახელება'].iloc[0]
            start_row = ws1.max_row + 1
            unique_invoices = group.groupby('სერია №')['ღირებულება დღგ და აქციზის ჩათვლით'].sum().reset_index()
            company_invoice_sum = unique_invoices['ღირებულება დღგ და აქციზის ჩათვლით'].sum()
            payment_formula = f"=SUMIF(საბანკოამონაწერი!P:P, B{start_row}, საბანკოამონაწერი!D:D)"

            ws1.append([company_name, company_id, '', company_invoice_sum, payment_formula])

            for _, row in unique_invoices.iterrows():
                ws1.append(['', '', row['სერია №'], row['ღირებულება დღგ და აქციზის ჩათვლით'], ''])

        ws2 = wb.create_sheet(title="დეტალური მონაცემები")
        ws2.append(purchases_df.columns.tolist())
        for row in purchases_df.itertuples(index=False):
            ws2.append(row)

        ws3 = wb.create_sheet(title="საბანკოამონაწერი")
        ws3.append(bank_df.columns.tolist())
        for row in bank_df.itertuples(index=False):
            ws3.append(row)

        ws4 = wb.create_sheet(title="ანგარიშფაქტურის დეტალები")
        invoice_details_df = purchases_df[['სერია №', 'საქონელი / მომსახურება', 'ზომის ერთეული', 'რაოდ.', 'ღირებულება დღგ და აქციზის ჩათვლით']].copy()
        invoice_details_df.rename(columns={'სერია №': 'ინვოისის №'}, inplace=True)
        ws4.append(invoice_details_df.columns.tolist())
        for row in invoice_details_df.itertuples(index=False):
            ws4.append(row)

        ws5 = wb.create_sheet(title="გადარიცხვები_უბმოლოდ")
        missing_payments = bank_df[~bank_df['P'].isin(purchases_df['საიდენტიფიკაციო კოდი'])]
        ws5.append(missing_payments.columns.tolist())
        for row in missing_payments.itertuples(index=False):
            ws5.append(row)

        ws6 = wb.create_sheet(title="განახლებული ამონაწერი")
        ws6.append(bank_df.columns.tolist())
        for row in bank_df.itertuples(index=False):
            ws6.append(row)

        ws7 = wb.create_sheet(title="კომპანიების_ჯამები")
        ws7.append(['დასახელება', 'საიდენტიფიკაციო კოდი', 'ანგარიშფაქტურების ჯამი', 'ჩარიცხული თანხა'])
        for row in range(2, ws1.max_row + 1):
            company_name = ws1[f"A{row}"].value
            company_id = ws1[f"B{row}"].value
            invoice_sum = ws1[f"D{row}"].value
            if company_name and company_id:
                payment_formula = f"=SUMIF(საბანკოამონაწერი!P:P, B{row}, საბანკოამონაწერი!D:D)"
                ws7.append([company_name, company_id, invoice_sum, payment_formula])

        final_file = "final_output.xlsx"
        wb.save(final_file)

        with open(final_file, "rb") as f:
            st.download_button(
                label="⬇ Download Final File",
                data=f,
                file_name=final_file,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
